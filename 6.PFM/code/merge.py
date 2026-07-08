"""
월별 카드 사용내역 엑셀(1월~5월) 통합 마이그레이션 스크립트

기능:
  ./6.PFM/data/user_card_history_2026_01.xlsx ~ _05.xlsx 를 읽어
  month 컬럼을 맨 앞에 추가하고 하나의 마스터 CSV로 병합한다.

실행 방법:
  1. 패키지 설치: pip install openpyxl
  2. 실행 명령어: python code/merge.py

출력:
  ./6.PFM/data/user_card_history_merged.csv (UTF-8 BOM)
  컬럼: month, Approved_Num, Approved_DateTime, Card_Name, Store_Name, Category, Amount_KRW
"""

import csv
import glob
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print(
        "[오류] openpyxl 라이브러리가 설치되어 있지 않습니다. "
        "터미널에서 'pip install openpyxl'을 실행한 후 다시 시도해주세요."
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# 경로 설정
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
OUTPUT_CSV = DATA_DIR / "user_card_history_merged.csv"

HEADER = [
    "month",
    "Approved_Num",
    "Approved_DateTime",
    "Card_Name",
    "Store_Name",
    "Category",
    "Amount_KRW",
]


def find_excel_files() -> list[Path]:
    """data 폴더에서 대상 엑셀 파일을 찾아 파일명 오름차순으로 정렬한다."""
    pattern = str(DATA_DIR / "user_card_history_2026_*.xlsx")
    files = [Path(p) for p in glob.glob(pattern)]
    files.sort(key=lambda p: p.name)
    return files


def extract_month(filename: str) -> str | None:
    """파일명에서 월 정보를 추출해 YYYY-MM 형태로 반환한다."""
    match = re.search(r"user_card_history_(\d{4})_(\d{2})", filename)
    if not match:
        return None
    year, month = match.group(1), match.group(2)
    return f"{year}-{month}"


def parse_excel_file(filepath: Path, month: str) -> list[list]:
    """단일 엑셀 파일의 활성 시트를 파싱해 유효한 데이터 행 리스트를 반환한다."""
    rows: list[list] = []
    try:
        # data_only=True: 하단 Total 행의 수식(=SUM) 대신 계산된 값을 읽음
        workbook = load_workbook(filepath, data_only=True)
        sheet = workbook.active
    except Exception as exc:  # noqa: BLE001
        print(f"  [경고] '{filepath.name}' 파일을 여는 중 오류 발생: {exc}")
        return rows

    skipped_total = 0
    skipped_empty = 0

    # 행 1은 헤더이므로 2행부터 순회
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 6:
            skipped_empty += 1
            continue

        approved_num = row[0]
        amount_krw = row[5]

        # Total 합계 행 제외 (A열이 'Total')
        if isinstance(approved_num, str) and approved_num.strip().lower() == "total":
            skipped_total += 1
            continue

        # 필수 값(승인번호, 금액)이 비어있으면 건너뜀
        if approved_num is None or amount_krw is None:
            skipped_empty += 1
            continue
        if str(approved_num).strip() == "" or str(amount_krw).strip() == "":
            skipped_empty += 1
            continue

        record = [
            month,
            str(approved_num).strip(),
            str(row[1]).strip() if row[1] is not None else "",
            str(row[2]).strip() if row[2] is not None else "",
            str(row[3]).strip() if row[3] is not None else "",
            str(row[4]).strip() if row[4] is not None else "",
            amount_krw,
        ]
        rows.append(record)

    print(
        f"  - 유효 데이터 {len(rows)}건 추출 "
        f"(Total 행 {skipped_total}건 / 빈 행 {skipped_empty}건 제외)"
    )
    return rows


def write_csv(all_rows: list[list], output_path: Path) -> None:
    """병합된 데이터를 UTF-8 BOM CSV로 저장한다."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADER)
        writer.writerows(all_rows)


def main() -> None:
    """엑셀 파일 탐색 → 파싱 → 병합 → CSV 저장 전체 파이프라인."""
    print("=" * 60)
    print("월별 카드 사용내역 엑셀 통합 시작")
    print("=" * 60)

    files = find_excel_files()
    if not files:
        print(f"[오류] 대상 엑셀 파일을 찾을 수 없습니다: {DATA_DIR}")
        print("       user_card_history_2026_01.xlsx ~ _05.xlsx 파일을 확인하세요.")
        sys.exit(1)

    print(f"[1단계] 대상 파일 {len(files)}개 발견 (파일명 오름차순 정렬 완료)")
    for f in files:
        print(f"  · {f.name}")

    all_rows: list[list] = []
    print("[2단계] 파일별 데이터 파싱 및 정제")
    for filepath in files:
        month = extract_month(filepath.name)
        if month is None:
            print(f"  [경고] '{filepath.name}'에서 월 정보를 추출하지 못해 건너뜁니다.")
            continue
        print(f"· {filepath.name} (month={month}) 처리 중...")
        rows = parse_excel_file(filepath, month)
        all_rows.extend(rows)

    if not all_rows:
        print("[오류] 병합할 유효 데이터가 없습니다.")
        sys.exit(1)

    print(f"[3단계] 총 {len(all_rows)}건 병합 완료")

    try:
        write_csv(all_rows, OUTPUT_CSV)
    except Exception as exc:  # noqa: BLE001
        print(f"[오류] CSV 저장 중 문제가 발생했습니다: {exc}")
        sys.exit(1)

    print(f"[4단계] CSV 저장 완료 (UTF-8 BOM): {OUTPUT_CSV}")
    print("=" * 60)
    print("모든 작업이 완료되었습니다.")
    print("=" * 60)


if __name__ == "__main__":
    main()
